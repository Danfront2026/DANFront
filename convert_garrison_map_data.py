"""
convert_garrison_map_data.py
Reads the updated Excel file and regenerates js/garrison-map-data.js.

Sheets processed: Budin, Temesvar, Yanik_and_Pápa, Eger, Kanizsa, Várad, Újvár, RumeliSilistre
Skipped:         Bosna, Kamaniçe, Rest of theRumeli
"""

import openpyxl
import json
import os
import re

EXCEL_PATH = os.path.expanduser(
    "~/Downloads/Dataset_Hungary and Rumili and Özi_new.xlsx"
)
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "js", "garrison-map-data.js")

# ---------------------------------------------------------------------------
# Sheet configs
# Each entry describes the column indices (0-based) for that sheet.
# ---------------------------------------------------------------------------
SHEET_CONFIGS = {
    "Budin": {
        "y": 0, "x": 1,
        "name_en": 2, "name_tr": 3, "name_local": 4,
        "type": 5,
        "p_en": 6, "p_tr": 7, "p_local": 8,
        "s_en": 9, "s_tr": 10, "s_local": 11,
        "date1": 12, "date2": 13, "date3": None, "date4": None,
        "country": 14, "size": 15,
        "notes": 16, "sources": 17,
        "source_sheet": "Budin",
    },
    "Temesvar": {
        "y": 0, "x": 1,
        "name_en": 2, "name_tr": 3, "name_local": 4,
        "type": 5,
        "p_en": 6, "p_tr": 7, "p_local": 8,
        "s_en": 9, "s_tr": 10, "s_local": 11,
        "date1": 12, "date2": 13, "date3": None, "date4": None,
        "country": 15, "size": 16,
        "notes": 17, "sources": 18,
        "source_sheet": "Temesvar",
    },
    "Yanik_and_Pápa": {
        "y": 0, "x": 1,
        "name_en": 2, "name_tr": 3, "name_local": 4,
        "type": 5,
        "p_en": 6, "p_tr": 7, "p_local": 8,
        "s_en": 9, "s_tr": 10, "s_local": 11,
        "date1": 12, "date2": 13, "date3": None, "date4": None,
        "country": 14, "size": 15,
        "notes": 16, "sources": 17,
        "source_sheet": "Yanik_and_Pápa",
    },
    "Eger": {
        "y": 0, "x": 1,
        "name_en": 2, "name_tr": 3, "name_local": 4,
        "type": 5,
        "p_en": 6, "p_tr": 7, "p_local": 8,
        "s_en": 9, "s_tr": 10, "s_local": 11,
        "date1": 12, "date2": 13, "date3": None, "date4": None,
        "country": 14, "size": 15,
        "notes": 16, "sources": 17,
        "source_sheet": "Eger",
    },
    "Kanizsa": {
        "y": 0, "x": 1,
        "name_en": 2, "name_tr": 3, "name_local": 4,
        "type": 5,
        "p_en": 6, "p_tr": 7, "p_local": 8,
        "s_en": 9, "s_tr": 10, "s_local": 11,
        "date1": 12, "date2": 13, "date3": None, "date4": None,
        "country": 14, "size": 15,
        "notes": 16, "sources": 17,
        "source_sheet": "Kanizsa",
    },
    "Várad": {
        "y": 0, "x": 1,
        "name_en": 2, "name_tr": 3, "name_local": 4,
        "type": 5,
        "p_en": 6, "p_tr": 7, "p_local": 8,
        "s_en": 9, "s_tr": 10, "s_local": 11,
        "date1": 12, "date2": 13, "date3": None, "date4": None,
        "country": 14, "size": 15,
        "notes": 16, "sources": 17,
        "source_sheet": "Várad",
    },
    "Újvár": {
        "y": 0, "x": 1,
        "name_en": 2, "name_tr": 3, "name_local": 4,
        "type": 5,
        "p_en": 6, "p_tr": 7, "p_local": 8,
        "s_en": 9, "s_tr": 10, "s_local": 11,
        "date1": 12, "date2": 13, "date3": None, "date4": None,
        "country": 14, "size": 15,
        "notes": 16, "sources": 17,
        "source_sheet": "Újvár",
    },
    "RumeliSilistre": {
        "y": 0, "x": 1,
        "name_en": 2, "name_tr": 3, "name_local": 4,
        "type": 5,
        "size": 6,
        "p_en": 7, "p_tr": 8, "p_local": 9,
        "s_en": 10, "s_tr": 11, "s_local": 12,
        "date1": 13, "date2": 14, "date3": 15, "date4": 16,
        "country": 17, "sources": 18, "notes": 19,
        "source_sheet": "RumeliSilistre",
    },
}


def normalize_date(val):
    """Convert en-dash dates to hyphen, strip whitespace."""
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace("\u2013", "-").replace("\u2014", "-")  # en-dash, em-dash
    return s


def get_cell(row, idx):
    """Safely get a cell value by 0-based column index."""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def str_val(row, idx):
    v = get_cell(row, idx)
    if v is None:
        return ""
    return str(v).strip()


def process_sheet(ws, cfg):
    garrisons = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip blank rows
        if not any(v is not None for v in row[:5]):
            continue

        lat_raw = get_cell(row, cfg["y"])
        lng_raw = get_cell(row, cfg["x"])
        try:
            lat = float(lat_raw)
            lng = float(lng_raw)
        except (TypeError, ValueError):
            print(f"  WARNING: skipping row with bad coordinates: {row[:5]}")
            continue

        name_en = str_val(row, cfg["name_en"])
        name_tr = str_val(row, cfg["name_tr"])
        name_local = str_val(row, cfg["name_local"])

        garrison_type = str_val(row, cfg["type"])

        p_en = str_val(row, cfg["p_en"])
        p_tr = str_val(row, cfg["p_tr"])
        p_local = str_val(row, cfg["p_local"])

        s_en = str_val(row, cfg["s_en"])
        s_tr = str_val(row, cfg["s_tr"])
        s_local = str_val(row, cfg["s_local"])

        date1 = normalize_date(get_cell(row, cfg["date1"]))
        date2 = normalize_date(get_cell(row, cfg.get("date2")))
        date3 = normalize_date(get_cell(row, cfg.get("date3")))
        date4 = normalize_date(get_cell(row, cfg.get("date4")))

        country = str_val(row, cfg["country"])
        size = str_val(row, cfg["size"])
        notes = str_val(row, cfg["notes"])
        sources = str_val(row, cfg["sources"])

        garrison = {
            "lat": round(lat, 6),
            "lng": round(lng, 6),
            "name": {
                "en": name_en,
                "tr": name_tr,
                "local": name_local,
            },
            "type": garrison_type,
            "province": {
                "en": p_en,
                "tr": p_tr,
                "local": p_local,
            },
            "sancak": {
                "en": s_en,
                "tr": s_tr,
                "local": s_local,
            },
            "date1": date1,
            "date2": date2,
            "date3": date3,
            "date4": date4,
            "country": country,
            "size": size,
            "notes": notes,
            "sources": sources,
            "source_sheet": cfg["source_sheet"],
        }
        garrisons.append(garrison)
    return garrisons


def garrison_to_js(g):
    """Render one garrison object as a JS object string (indented 4 spaces)."""
    lines = ["    {"]
    lines.append(f'      "lat": {g["lat"]},')
    lines.append(f'      "lng": {g["lng"]},')
    lines.append(f'      "name": {{')
    lines.append(f'        "en": {json.dumps(g["name"]["en"])},')
    lines.append(f'        "tr": {json.dumps(g["name"]["tr"])},')
    lines.append(f'        "local": {json.dumps(g["name"]["local"])}')
    lines.append(f'      }},')
    lines.append(f'      "type": {json.dumps(g["type"])},')
    lines.append(f'      "province": {{')
    lines.append(f'        "en": {json.dumps(g["province"]["en"])},')
    lines.append(f'        "tr": {json.dumps(g["province"]["tr"])},')
    lines.append(f'        "local": {json.dumps(g["province"]["local"])}')
    lines.append(f'      }},')
    lines.append(f'      "sancak": {{')
    lines.append(f'        "en": {json.dumps(g["sancak"]["en"])},')
    lines.append(f'        "tr": {json.dumps(g["sancak"]["tr"])},')
    lines.append(f'        "local": {json.dumps(g["sancak"]["local"])}')
    lines.append(f'      }},')
    lines.append(f'      "date1": {json.dumps(g["date1"])},')
    lines.append(f'      "date2": {json.dumps(g["date2"])},')
    lines.append(f'      "date3": {json.dumps(g["date3"])},')
    lines.append(f'      "date4": {json.dumps(g["date4"])},')
    lines.append(f'      "country": {json.dumps(g["country"])},')
    lines.append(f'      "size": {json.dumps(g["size"])},')
    lines.append(f'      "notes": {json.dumps(g["notes"])},')
    lines.append(f'      "sources": {json.dumps(g["sources"])},')
    lines.append(f'      "source_sheet": {json.dumps(g["source_sheet"])}')
    lines.append("    }")
    return "\n".join(lines)


def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_garrisons = []
    for sheet_name, cfg in SHEET_CONFIGS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  SHEET NOT FOUND: {sheet_name} — skipping")
            continue
        ws = wb[sheet_name]
        garrisons = process_sheet(ws, cfg)
        print(f"  {sheet_name}: {len(garrisons)} garrisons")
        all_garrisons.extend(garrisons)

    total = len(all_garrisons)
    print(f"\nTotal garrisons: {total}")

    # Build JS content
    garrison_js_list = []
    for i, g in enumerate(all_garrisons):
        s = garrison_to_js(g)
        if i < total - 1:
            s += ","
        garrison_js_list.append(s)

    garrison_block = "\n".join(garrison_js_list)

    js_content = f"""/**
 * Garrison Map Data
 * Coordinates and metadata for map display.
 * Generated automatically by convert_garrison_map_data.py - do not edit manually.
 *
 * Schema per garrison:
 *   lat, lng          — coordinates
 *   name              — {{en, tr, local}}  (local = HU for Hungary sheets, SRB/BG/RO/UKR for Rumeli)
 *   type              — Fortress | Palanka | Parkan | ...
 *   province          — {{en, tr, local}}
 *   sancak            — {{en, tr, local}}
 *   date1..date4      — Ottoman occupation periods (en-dash normalized to hyphen)
 *   country           — modern country name
 *   size              — Military Hub | Large | Medium | Small | ""
 *   notes             — free-text annotation
 *   sources           — abbreviated source citation(s)
 *   source_sheet      — Excel sheet of origin
 */

var garrisonMapData = {{
  "total_garrisons": {total},
  "garrisons": [
{garrison_block}
  ]
}};
"""

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"\nWrote {OUTPUT_PATH}")
    print("Done.")


if __name__ == "__main__":
    main()
